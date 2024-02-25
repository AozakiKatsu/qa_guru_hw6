import csv
import os
import shutil
import zipfile
import pytest
from openpyxl import load_workbook
from pypdf import PdfReader
from io import TextIOWrapper

# настраиваем пути
dir_path = os.path.join(os.getcwd(), 'tmp')
file_dir = os.listdir(dir_path)
zip_path = os.path.join(os.getcwd(), 'resource')


@pytest.fixture(scope='function', autouse=True)
def create_archive():
    if not os.path.exists(zip_path):
        os.mkdir(zip_path)
        with zipfile.ZipFile('docs_archive.zip', 'w') as zf:
            for file in file_dir:
                add_file = os.path.join(dir_path, file)
                zf.write(add_file, os.path.basename(add_file))
    shutil.move(os.path.abspath('docs_archive.zip'), zip_path)
    yield
    shutil.rmtree(zip_path)


def test_pdf():
    with zipfile.ZipFile('resource/docs_archive.zip', "r") as zf:
        with zf.open('Pdf.pdf', "r") as text1:
            reader = PdfReader(text1)

            assert 'Python Testing with pytest' in reader.pages[1].extract_text()
            print('OK pdf')


def test_xlsx():
    with zipfile.ZipFile('resource/docs_archive.zip', "r") as zf:
        with zf.open('Xlsx.xlsx', "r") as text2:
            workbook = load_workbook(text2)
            sheet = workbook.active

            first_assert = sheet.cell(row=4, column=2).value
            assert first_assert == 'Philip'
            print('OK xlsx')


def test_csv():
    with zipfile.ZipFile('resource/docs_archive.zip', "r") as zf:
        with zf.open('Csv.csv', "r") as text3:
            csvreader = list(csv.reader(TextIOWrapper(text3, 'windows-1251'), delimiter=';'))
            second_row = csvreader[1]

            assert second_row[0] == 'CN001'
            assert second_row[1] == 'Иванова'
            print('OK xlsx')
